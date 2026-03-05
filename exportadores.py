import pandas as pd
import json
from collections import Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colores
COLOR_HEADER      = "1565C0"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_FILA_PAR    = "E3F2FD"
COLOR_FILA_IMPAR  = "FFFFFF"
COLOR_COMPLETADA  = "C8E6C9"
COLOR_EN_PROGRESO = "FFF9C4"
COLOR_SUSPENDIDA  = "FFCCBC"
COLOR_EN_SITIO    = "B3E5FC"
COLOR_PENDIENTE   = "FFE0B2"

BORDE = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
BORDE_MEDIO = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium")
)


def celda(ws, fila, col, valor, bold=False, size=11, color_fondo=None,
          color_font="000000", alineacion="left", wrap=False, borde=BORDE):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font      = Font(bold=bold, size=size, color=color_font)
    c.alignment = Alignment(horizontal=alineacion, vertical="center", wrap_text=wrap)
    if color_fondo:
        c.fill = PatternFill("solid", fgColor=color_fondo)
    if borde:
        c.border = borde
    return c


def aplicar_formato(ws, df, color_header=COLOR_HEADER, freeze=True):
    """Aplica formato profesional a una hoja de Excel."""
    header_fill = PatternFill("solid", fgColor=color_header)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    fila_par    = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    fila_impar  = PatternFill("solid", fgColor=COLOR_FILA_IMPAR)

    # Headers
    for col_idx in range(1, len(df.columns) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas alternas
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        fill = fila_par if row_idx % 2 == 0 else fila_impar
        for c in row:
            c.fill = fill
            c.border = BORDE
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # Anchos de columna dinámicos
    for col_idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col))
        for val in df[col].astype(str):
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[col_letter].width = max_len + 4

    ws.row_dimensions[1].height = 30
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 40

    ws.auto_filter.ref = ws.dimensions
    if freeze:
        ws.freeze_panes = "A2"


def aplicar_condicional_estado(ws, df):
    """Colorea filas según el estado de la actividad."""
    if "ESTADO" not in df.columns:
        return
    
    estado_col_idx = list(df.columns).index("ESTADO") + 1
    
    colores = {
        "COMPLETADA": PatternFill("solid", fgColor=COLOR_COMPLETADA),
        "EN PROGRESO": PatternFill("solid", fgColor=COLOR_EN_PROGRESO),
        "SUSPENDIDA": PatternFill("solid", fgColor=COLOR_SUSPENDIDA),
        "EN SITIO": PatternFill("solid", fgColor=COLOR_EN_SITIO),
        "PENDIENTE AUTORIZACIÓN": PatternFill("solid", fgColor=COLOR_PENDIENTE),
    }
    
    for row_idx in range(2, ws.max_row + 1):
        val = str(ws.cell(row=row_idx, column=estado_col_idx).value).strip().upper()
        fill = colores.get(val)
        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill


def construir_resumen(ws, df_actividades, stats):
    """Construye hoja Resumen estilo dashboard."""
    
    for i, w in enumerate([4, 28, 20, 28, 20, 4], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in range(1, 100):
        ws.row_dimensions[row].height = 22

    fila = 2

    # TÍTULO
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=5)
    celda(ws, fila, 2, "RESUMEN DE ACTIVIDADES - INSTALACIONES",
          bold=True, size=16, color_fondo="1565C0", color_font="FFFFFF",
          alineacion="center")
    ws.row_dimensions[fila].height = 35
    fila += 2

    # KPIs
    total = stats.get("total_actividades", 0)
    completadas = stats.get("por_estado", {}).get("Completada", 0)
    en_progreso = stats.get("por_estado", {}).get("En Progreso", 0)
    suspendidas = stats.get("por_estado", {}).get("Suspendida", 0)

    kpis = [
        ("ACTIVIDADES TOTALES", str(total), "1565C0"),
        ("COMPLETADAS", str(completadas), "2E7D32"),
        ("EN PROGRESO", str(en_progreso), "F9A825"),
        ("SUSPENDIDAS", str(suspendidas), "E65100"),
    ]
    
    col_kpi = 2
    for titulo, valor, color in kpis:
        ws.merge_cells(start_row=fila, start_column=col_kpi, end_row=fila, end_column=col_kpi)
        celda(ws, fila, col_kpi, titulo, bold=True, size=9,
              color_fondo=color, color_font="FFFFFF", alineacion="center", borde=BORDE_MEDIO)
        ws.merge_cells(start_row=fila+1, start_column=col_kpi, end_row=fila+1, end_column=col_kpi)
        celda(ws, fila+1, col_kpi, valor, bold=True, size=22,
              color_fondo="F5F5F5", alineacion="center", borde=BORDE_MEDIO)
        ws.row_dimensions[fila+1].height = 45
        col_kpi += 1
    fila += 3

    def bloque_tabla(ws, fila, col_ini, titulo, color_hdr, headers, filas_data):
        ws.merge_cells(start_row=fila, start_column=col_ini,
                       end_row=fila, end_column=col_ini + len(headers) - 1)
        celda(ws, fila, col_ini, titulo, bold=True, size=11,
              color_fondo=color_hdr, color_font="FFFFFF", alineacion="center")
        fila += 1
        for i, h in enumerate(headers):
            celda(ws, fila, col_ini + i, h, bold=True, size=10,
                  color_fondo="BBDEFB", alineacion="center")
        fila += 1
        for idx, row_data in enumerate(filas_data):
            bg = COLOR_FILA_PAR if idx % 2 == 0 else COLOR_FILA_IMPAR
            for i, val in enumerate(row_data):
                celda(ws, fila, col_ini + i, val, size=10, color_fondo=bg, alineacion="center")
            fila += 1
        return fila

    if not df_actividades.empty:
        # Top Clientes | Top Técnicos
        top_cli = list(stats.get("top_clientes", {}).items())[:8]
        top_tec = list(stats.get("top_tecnicos", {}).items())[:8]
        
        fila_ini = fila
        fila_a = bloque_tabla(ws, fila_ini, 2, "TOP CLIENTES CON MAS ACTIVIDADES", "1565C0",
            ["Cliente", "Actividades"],
            [(cli, cant) for cli, cant in top_cli])
        fila_b = bloque_tabla(ws, fila_ini, 4, "TOP TÉCNICOS MÁS PRODUCTIVOS", "C62828",
            ["Técnico", "Actividades"],
            [(tec, cant) for tec, cant in top_tec])
        fila = max(fila_a, fila_b) + 1

        # Distribución por tipo | Por ciudad
        por_tipo = list(stats.get("por_tipo", {}).items())[:8]
        por_ciudad = list(stats.get("por_ciudad", {}).items())[:8]
        
        fila_ini2 = fila
        fila_c = bloque_tabla(ws, fila_ini2, 2, "DISTRIBUCIÓN POR TIPO DE ACTIVIDAD", "6A1B9A",
            ["Tipo", "Cantidad"],
            [(tipo, cant) for tipo, cant in por_tipo])
        fila_d = bloque_tabla(ws, fila_ini2, 4, "ACTIVIDADES POR CIUDAD", "00695C",
            ["Ciudad", "Cantidad"],
            [(ciudad, cant) for ciudad, cant in por_ciudad])
        fila = max(fila_c, fila_d) + 1


def generar_archivos(df_actividades, df_chat, stats, excel_path, json_path, csv_path):
    """Genera archivos de salida: Excel, JSON, CSV."""
    
    # Excel con múltiples hojas
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        if not df_actividades.empty:
            df_actividades.to_excel(writer, sheet_name="Actividades", index=False)
        else:
            pd.DataFrame([{"Mensaje": "Sin actividades"}]).to_excel(writer, sheet_name="Actividades", index=False)
        
        if not df_chat.empty:
            df_chat.to_excel(writer, sheet_name="Chat_WhatsApp", index=False)
        
        pd.DataFrame().to_excel(writer, sheet_name="Resumen", index=False)
    
    # Aplicar formato
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    
    # Formato Actividades
    ws_act = wb["Actividades"]
    df_fmt = df_actividades if not df_actividades.empty else pd.DataFrame([{"Mensaje": "Sin actividades"}])
    aplicar_formato(ws_act, df_fmt)
    if not df_actividades.empty:
        aplicar_condicional_estado(ws_act, df_actividades)
    
    # Formato Chat
    if "Chat_WhatsApp" in wb.sheetnames and not df_chat.empty:
        aplicar_formato(wb["Chat_WhatsApp"], df_chat, color_header="37474F")
    
    # Dashboard Resumen
    ws_res = wb["Resumen"]
    construir_resumen(ws_res, df_actividades, stats)
    
    wb.save(excel_path)
    
    # JSON
    records = df_actividades.to_dict("records") if not df_actividades.empty else []
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    
    # CSV
    if not df_actividades.empty:
        df_actividades.to_csv(csv_path, index=False, encoding="utf-8-sig")
